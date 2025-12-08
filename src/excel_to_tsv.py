#!/usr/bin/env python3
"""
Convert Excel file to TSV format
Extracts INDEX (VISA), TITLE, REALISATEUR, DEVIS, and DATE columns from all year sheets
"""

import sys
import csv

def excel_to_tsv(excel_path, tsv_path):
    """
    Convert Excel file to TSV format
    Reads all year sheets (2003-2024) and combines them
    Args:
        excel_path: Path to the Excel file
        tsv_path: Path to output TSV file
    """
    try:
        from openpyxl import load_workbook
        
        # Load the workbook
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        
        # Get all year sheets (sheets with numeric names)
        year_sheets = [name for name in wb.sheetnames if name.isdigit()]
        year_sheets.sort(reverse=True)  # Sort from newest to oldest
        
        print(f"Found {len(year_sheets)} year sheets: {', '.join(year_sheets)}")
        
        # Write to TSV file
        with open(tsv_path, 'w', newline='', encoding='utf-8') as tsvfile:
            writer = csv.writer(tsvfile, delimiter='\t')
            
            # Write header
            writer.writerow(['INDEX', 'TITRE', 'REALISATEUR', 'DEVIS', 'DATE'])
            
            total_rows = 0
            
            # Process each year sheet
            for year in year_sheets:
                ws = wb[year]
                
                # Find the header row (contains 'VISA', 'TITRE', etc.)
                header_row = None
                visa_idx = None
                titre_idx = None
                realisateur_idx = None
                devis_idx = None
                
                for row_num, row in enumerate(ws.iter_rows(values_only=True), 1):
                    if row and any(cell for cell in row):
                        # Check if this row contains headers
                        row_str = [str(cell).upper() if cell else '' for cell in row]
                        if 'VISA' in row_str or 'TITRE' in row_str:
                            header_row = row_num
                            for idx, cell in enumerate(row_str):
                                if 'VISA' in cell:
                                    visa_idx = idx
                                elif 'TITRE' in cell:
                                    titre_idx = idx
                                elif 'RÉALISATEUR' in cell or 'REALISATEUR' in cell:
                                    realisateur_idx = idx
                                elif 'DEVIS' in cell:
                                    devis_idx = idx
                            break
                
                if header_row is None:
                    print(f"Warning: Could not find header in sheet {year}")
                    continue
                
                # Read data rows (skip header)
                sheet_rows = 0
                for row in ws.iter_rows(min_row=header_row+1, values_only=True):
                    if not row or not any(cell for cell in row):
                        continue
                    
                    # Extract values
                    visa_val = row[visa_idx] if visa_idx is not None and visa_idx < len(row) else ''
                    titre_val = row[titre_idx] if titre_idx is not None and titre_idx < len(row) else ''
                    realisateur_val = row[realisateur_idx] if realisateur_idx is not None and realisateur_idx < len(row) else ''
                    devis_val = row[devis_idx] if devis_idx is not None and devis_idx < len(row) else ''
                    
                    # Skip rows with empty VISA (index)
                    if not visa_val or visa_val == '':
                        continue
                    
                    # Convert None to empty string, keep numbers and strings as-is
                    visa_val = '' if visa_val is None else str(visa_val)
                    titre_val = '' if titre_val is None else str(titre_val)
                    realisateur_val = '' if realisateur_val is None else str(realisateur_val)
                    devis_val = '' if devis_val is None else str(devis_val)
                    
                    # Write row with year as DATE
                    writer.writerow([visa_val, titre_val, realisateur_val, devis_val, year])
                    sheet_rows += 1
                
                total_rows += sheet_rows
                print(f"  Processed {sheet_rows} rows from sheet {year}")
        
        wb.close()
        print(f"Successfully converted {excel_path} to {tsv_path}")
        print(f"Total rows written: {total_rows}")
        return True
        
    except ImportError:
        print("Error: openpyxl not installed. Install with: pip install openpyxl")
        print("Alternatively, try: python -m venv venv && source venv/bin/activate && pip install openpyxl")
        return False
    except Exception as e:
        print(f"Error converting Excel to TSV: {e}")
        import traceback
        traceback.print_exc()
        return False


if __name__ == "__main__":
    excel_file = "data/production cinématographique - liste des premiers films.xlsx"
    tsv_file = "data/parsed/productionCinématographique.tsv"
    
    if len(sys.argv) > 1:
        excel_file = sys.argv[1]
    if len(sys.argv) > 2:
        tsv_file = sys.argv[2]
    
    success = excel_to_tsv(excel_file, tsv_file)
    sys.exit(0 if success else 1)
